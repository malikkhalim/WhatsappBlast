#Author @malikkhalim
#Use at your own risk
#The script may not work in case if the HTML of web WhatsApp is changed. Keep updating the code. If you face any problem please do let me know. Surely I will help you out.
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
import pandas as pd
import logging
import openpyxl
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Configure logging
logging.basicConfig(filename='error.log', level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

# Create an empty DataFrame to store the results
results = pd.DataFrame(columns=['Name', 'Status'])

excel_data = pd.read_excel('Recipients data.xlsx', sheet_name='Recipients')

count = 0

Tk().withdraw()  # Hide the main tkinter window
image_file_path = askopenfilename(title='Select Image File', filetypes=[('Image Files', '*.jpeg;*.jpg;*.png')])


webdriver_service = Service(ChromeDriverManager().install())

# Configure the Chrome options if needed
chrome_options = webdriver.ChromeOptions()
# Add any necessary options to chrome_options

# Create the WebDriver with the specified service and options
driver = webdriver.Chrome(service=webdriver_service, options=chrome_options)
driver.get('https://web.whatsapp.com')
input("Press ENTER after login into WhatsApp Web and your chats are visible.")

for column in excel_data['Contact'].tolist():
    try:
        name = excel_data['Name'][count]  # Extract the name from the Excel sheet
        message = excel_data['Message'][0].replace('{name}', name)  # Replace '{name}' with the recipient's name

        # Generate the URL with the modified message
        url = 'https://web.whatsapp.com/send?phone=62{}&text={}'.format(excel_data['Contact'][count], message)
        driver.get(url)

        sleep(8)

        # Check if the error alert is present on the page
        error_alert = driver.find_elements(By.XPATH, '//div[contains(@class, "iuhl9who") and text()="Phone number shared via url is invalid."]')
        if len(error_alert) > 0:
            raise Exception('Invalid URL number')

        # Check if the attachment button is visible
        attachment_btn = WebDriverWait(driver, 35).until(EC.element_to_be_clickable((By.CLASS_NAME, '_1OT67')))
        attachment_btn.click()

        sleep(1)
        

        # Set the image file path
        # image_file_path = "C:/Users/akbar/Iti_brosur.jpeg"  # Replace with the correct image file path

        # Locate the file input element and send the file path
        file_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[1]/div/div/span/div/ul/div/div[1]/li/div/input')))
        file_input.send_keys(image_file_path)

        sleep(1)

        # Wait for the send button to be clickable
        send_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, '_3wFFT')))
        send_btn.click()

        sleep(5)
        print('Message sent to: ' + str(excel_data['Name'][count]))

        # Update the result directly to the Excel file
        filename = 'Results.xlsx'

        book = load_workbook(filename)
        sheet_name = book.sheetnames[0]
        sheet = book[sheet_name]

        sheet.cell(row=count + 2, column=1).value = excel_data['Name'][count]
        sheet.cell(row=count + 2, column=2).value = 'Terkirim Info Pendaftaran'

        book.save(filename)

        count += 1
    except Exception as e:
        error_msg = 'Failed to send a message to {}: {}'.format(str(excel_data['Name'][count]), str(e))
        print(error_msg)
        logging.error(error_msg)

        # Update the result directly in the Excel file
        filename = 'Results.xlsx'

        book = load_workbook(filename)
        sheet_name = book.sheetnames[0]
        sheet = book[sheet_name]

        sheet.cell(row=count + 2, column=1).value = excel_data['Name'][count]
        sheet.cell(row=count + 2, column=2).value = 'Failed'

        book.save(filename)

        count += 1

driver.quit()
print("The script executed successfully.")


